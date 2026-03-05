from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import smtplib
import imaplib
import email
from email.message import EmailMessage
from email.header import decode_header
import threading
import time
import re
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)

EXCEL_FILE = "orders.xlsx"

# -------------------------------------------------
# Excel initialisieren
# -------------------------------------------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Zeit",
            "Bestellnummer",
            "Anrede", "Vorname", "Nachname",
            "E-Mail", "Telefon",
            "Straße", "Hausnr", "PLZ", "Ort",
            "Menge (Liter)", "Tankart", "Einfüllstutzen",
            "Bemerkung",
            "Status",
            "Bestätigung am"
        ])
        wb.save(EXCEL_FILE)

def save_to_excel(order, order_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        order_id,
        order.get("anrede"),
        order.get("vorname"),
        order.get("nachname"),
        order.get("email"),
        order.get("telefon"),
        order.get("strasse"),
        order.get("hausnr"),
        order.get("plz"),
        order.get("ort"),
        order.get("menge"),
        order.get("tankart"),
        order.get("einfuellstutzen"),
        order.get("bemerkung", ""),
        "WARTET AUF BESTÄTIGUNG",
        ""   # Bestätigung am – wird später gefüllt
    ])
    wb.save(EXCEL_FILE)

def update_excel_status(order_id: str, new_status: str):
    """Setzt Status und Bestätigungsdatum für eine Bestellnummer."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Spaltenindizes (1-basiert) ermitteln
    header = {cell.value: cell.column for cell in ws[1]}
    col_id     = header.get("Bestellnummer")
    col_status = header.get("Status")
    col_date   = header.get("Bestätigung am")

    for row in ws.iter_rows(min_row=2):
        if row[col_id - 1].value == order_id:
            row[col_status - 1].value = new_status
            if col_date:
                row[col_date - 1].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break

    wb.save(EXCEL_FILE)
    print(f"[Excel] Status für {order_id} → {new_status}")

# -------------------------------------------------
# E-Mail senden via iCloud SMTP
# -------------------------------------------------
def send_confirmation_mail(order: dict, order_id: str):
    """
    Sendet die Bestätigungsmail über das iCloud-Konto heizoel@beierhq.de.
    iCloud-SMTP: smtp.mail.me.com : 587 (STARTTLS)
    """
    mail_user = os.getenv("MAIL_USER")   # heizoel@beierhq.de
    mail_pass = os.getenv("MAIL_PASS")   # App-spezifisches iCloud-Passwort

    msg = EmailMessage()
    msg["From"]    = mail_user
    msg["To"]      = order["email"]
    msg["Subject"] = f"Heizölbestellung – Bitte bestätigen (Nr. {order_id})"
    # Reply-To auf dieselbe Adresse, damit Antworten im iCloud-Postfach landen
    msg["Reply-To"] = mail_user

    msg.set_content(f"""\
Sehr geehrte(r) {order.get("anrede", "")} {order.get("vorname", "")} {order.get("nachname", "")},

vielen Dank für Ihre Heizölbestellung.
Bitte prüfen Sie die Angaben und bestätigen Sie, indem Sie auf diese E-Mail antworten
und das Wort BESTÄTIGEN in den Text schreiben.

────────────────────────────
🧾 BESTELLDETAILS
────────────────────────────
Bestellnummer:    {order_id}

Menge:            {order.get("menge")} Liter
Tankart:          {order.get("tankart")}
Einfüllstutzen:   {order.get("einfuellstutzen")}

────────────────────────────
📍 LIEFERADRESSE
────────────────────────────
{order.get("anrede")} {order.get("vorname")} {order.get("nachname")}
{order.get("strasse")} {order.get("hausnr")}
{order.get("plz")} {order.get("ort")}

────────────────────────────
📞 KONTAKT
────────────────────────────
E-Mail:   {order.get("email")}
Telefon:  {order.get("telefon")}

────────────────────────────
📝 BEMERKUNG
────────────────────────────
{order.get("bemerkung", "—")}

Ohne Bestätigung wird die Bestellung nicht ausgeführt.

Mit freundlichen Grüßen
Ihr Heizöl-Service
""")

    # iCloud nutzt STARTTLS auf Port 587
    with smtplib.SMTP("smtp.mail.me.com", 587) as server:
        server.ehlo()
        server.starttls()
        server.login(mail_user, mail_pass)
        server.send_message(msg)

    print(f"[Mail] Bestätigungsmail gesendet an {order['email']} (Order: {order_id})")

# -------------------------------------------------
# IMAP-Polling – Antworten aus dem Ordner "heizoel"
# -------------------------------------------------
IMAP_SERVER  = "imap.mail.me.com"
IMAP_PORT    = 993
IMAP_FOLDER  = "heizoel"          # iCloud-Ordner, in den Antworten einsortiert werden
POLL_INTERVAL = 120               # Sekunden zwischen den Abfragen

# Bestellnummer aus Betreff oder Body extrahieren
ORDER_ID_RE = re.compile(r"BO-\d{14}", re.IGNORECASE)

def _decode_header_value(raw) -> str:
    parts = decode_header(raw or "")
    decoded = []
    for part, enc in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(enc or "utf-8", errors="replace"))
        else:
            decoded.append(part)
    return " ".join(decoded)

def _get_body(msg) -> str:
    """Gibt den Plain-Text-Inhalt einer E-Mail zurück."""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                return part.get_payload(decode=True).decode(
                    part.get_content_charset() or "utf-8", errors="replace"
                )
    else:
        return msg.get_payload(decode=True).decode(
            msg.get_content_charset() or "utf-8", errors="replace"
        )
    return ""

def poll_imap():
    """
    Läuft als Hintergrund-Thread.
    Verbindet sich alle POLL_INTERVAL Sekunden per IMAP mit iCloud,
    liest ungelesene Nachrichten aus dem Ordner "heizoel" und
    aktualisiert den Excel-Status, wenn "BESTÄTIGEN" im Body steht.
    """
    mail_user = os.getenv("MAIL_USER")
    mail_pass = os.getenv("MAIL_PASS")

    while True:
        try:
            with imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT) as imap:
                imap.login(mail_user, mail_pass)

                # Ordner auswählen
                status, _ = imap.select(f'"{IMAP_FOLDER}"')
                if status != "OK":
                    print(f"[IMAP] Ordner '{IMAP_FOLDER}' nicht gefunden.")
                    time.sleep(POLL_INTERVAL)
                    continue

                # Nur ungelesene Nachrichten
                _, msg_ids = imap.search(None, "UNSEEN")
                ids = msg_ids[0].split()

                if ids:
                    print(f"[IMAP] {len(ids)} neue Nachricht(en) im Ordner '{IMAP_FOLDER}'")

                for mid in ids:
                    _, data = imap.fetch(mid, "(RFC822)")
                    raw = data[0][1]
                    msg = email.message_from_bytes(raw)

                    subject = _decode_header_value(msg.get("Subject", ""))
                    body    = _get_body(msg)

                    # Bestellnummer suchen – zuerst im Betreff, dann im Body
                    match = ORDER_ID_RE.search(subject) or ORDER_ID_RE.search(body)

                    if match:
                        order_id = match.group(0).upper()
                        # Prüfen ob "BESTÄTIGEN" (case-insensitive) im Body vorkommt
                        if re.search(r"best[äa]tig", body, re.IGNORECASE):
                            update_excel_status(order_id, "BESTÄTIGT")
                            # Nachricht als gelesen markieren
                            imap.store(mid, "+FLAGS", "\\Seen")
                            print(f"[IMAP] Bestellung {order_id} bestätigt.")
                        else:
                            print(f"[IMAP] Antwort für {order_id} ohne Bestätigungswort.")
                    else:
                        print(f"[IMAP] Keine Bestellnummer in Nachricht gefunden: {subject!r}")

        except Exception as e:
            print(f"[IMAP] Fehler beim Polling: {e}")

        time.sleep(POLL_INTERVAL)

# -------------------------------------------------
# Routes
# -------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    order    = request.json
    order_id = f"BO-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    save_to_excel(order, order_id)

    try:
        send_confirmation_mail(order, order_id)
    except Exception as e:
        print(f"[Mail] Fehler beim Senden: {e}")
        return jsonify({"status": "mail_error", "order_id": order_id, "error": str(e)}), 500

    return jsonify({"status": "ok", "order_id": order_id})

# -------------------------------------------------
if __name__ == "__main__":
    init_excel()

    # IMAP-Polling als Daemon-Thread starten
    poller = threading.Thread(target=poll_imap, daemon=True)
    poller.start()

    app.run(debug=True)